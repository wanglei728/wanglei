"""
@Time : 2021/6/10 15:46
@Author : 
"""
import openpyxl


class HandlerExcel:
    """封装excel数据读取和写入方法"""

    def __init__(self, filename, sheetname):
        """初始化excel对象属性，文件名称和表单名称"""
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        # 1、创建excel工作簿和表单对象
        workbook = openpyxl.load_workbook(filename=self.filename)
        sheet = workbook[self.sheetname]
        # 2、：按行读取excel所有数据并转化为list
        excel_data = list(sheet.rows)
        # 3、：excel数据转化为测试用例数据格式（list中元素是dict格式）
        cases = []
        list_1 = [c.value for c in excel_data[0]]
        for item in excel_data[1:]:
            list_x = [c.value for c in item]
            dic = dict(zip(list_1, list_x))
            cases.append(dic)
        return cases

    def write_excel(self, row, column, value):
        workbook = openpyxl.load_workbook(filename=self.filename)
        sheet = workbook[self.sheetname]
        sheet.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


if __name__ == '__main__':
    excel = HandlerExcel(filename=r"D:\PyCharm\hetproject\project_het\datas\cases_data.xlsx", sheetname="app")
    res = excel.read_excel()
    print(res)
